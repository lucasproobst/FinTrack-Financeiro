from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


def gerar_relatorio_pdf(context):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(
        Paragraph(
            f"<b>Relatório Financeiro - {context['data_inicio']} a {context['data_fim']}</b>",
            styles["Title"],
        )
    )
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(f"Saldo Inicial: {context['saldo_inicial_fmt']}", styles["Normal"])
    )
    elements.append(
        Paragraph(f"Total de Receitas: {context['receitas_fmt']}", styles["Normal"])
    )
    elements.append(
        Paragraph(f"Total de Despesas: {context['despesas_fmt']}", styles["Normal"])
    )
    elements.append(
        Paragraph(f"<b>Saldo Final: {context['saldo_final_fmt']}</b>", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    data = [["Data", "Tipo", "Categoria", "Conta", "Valor", "Descrição"]]

    for l in context["lancamentos"]:
        # 🔥 Aceita ambos os padrões: "R"/"D" ou "Receita"/"Despesa"
        if l.tipo in ["R", "Receita"]:
            tipo = "Receita"
        else:
            tipo = "Despesa"

        categoria = l.categoria.nome if l.categoria else "-"
        conta = l.conta.nome if l.conta else "-"

        valor_formatado = (
            f"R$ {l.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )

        data.append(
            [
                l.data.strftime("%d/%m/%Y"),
                tipo,
                categoria,
                conta,
                valor_formatado,
                l.descricao,
            ]
        )

    table = Table(data, colWidths=[60, 60, 90, 80, 100, 130])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def gerar_relatorio_excel(context):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Financeiro"

    # Resumo
    ws.append(["Relatório", f"{context['data_inicio']} a {context['data_fim']}"])
    ws.append(["Saldo Inicial", context["saldo_inicial_fmt"]])
    ws.append(["Receitas", context["receitas_fmt"]])
    ws.append(["Despesas", context["despesas_fmt"]])
    ws.append(["Saldo Final", context["saldo_final_fmt"]])
    ws.append([])

    # Cabeçalho da tabela
    header = ["Data", "Tipo", "Categoria", "Conta", "Valor", "Descrição"]
    ws.append(header)

    for col in ws[7]:
        col.font = Font(bold=True)

    # Linhas da tabela
    for l in context["lancamentos"]:
        tipo = l.tipo  # CORREÇÃO FINAL

        categoria = l.categoria.nome if l.categoria else "-"
        conta = l.conta.nome if l.conta else "-"

        valor_formatado = (
            f"R$ {l.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )

        ws.append(
            [
                l.data.strftime("%d/%m/%Y"),
                tipo,
                categoria,
                conta,
                valor_formatado,
                l.descricao,
            ]
        )

    ws.column_dimensions["E"].width = 20

    for row in ws.iter_rows(min_row=8, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
